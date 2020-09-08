from argparse import ArgumentParser
import matplotlib.pyplot as plt
from numpy import histogram,max,argmax,nan,isnan
from pandas import read_csv
from pandas import concat
from cavefinomics import AstyanaxLi
from openpyxl.workbook import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter
import json

parser = ArgumentParser(description="Compute Inchi keys for compounds.")
parser.add_argument("--outlier", type=str, help="Use outlier or not.")
parser.add_argument("--level", type=float, help="P value cutoff.")
parser.add_argument('--output', type=str, help="Output file.")
args = parser.parse_args()



pops = ['Pachon', 'Tinaja', 'Surface']
tissues = ['Brain', 'Muscle', 'Liver']
conditions = ['4d Starved', '30d Starved', 'Refed']
conditions_really_short = ['4', r'30', 'R']
condmap = {'30d':'30d Starved', '4d':'4d Starved', 'Ref': 'Refed'}
#comparisons = ['PvS','TvS','PvT']
comparisons = {'PvS':('Pachon','Surface'), 'TvS':('Tinaja','Surface'), 'PvT':('Pachon','Tinaja')}
categories = {"Aminoacids":'Amino acids',"Carbohydrates_-CCM": 'Carbohydrates / CCM',"Fattyacids":'Fatty acids',"Misc._-_sec.metabolites":'Misc',"Nucleotides":'Nucleotides'}



datasets = []
sig = {}
up = {}
for cat in categories:
    for tissue in tissues:
        for cond in condmap:
            for comp in comparisons:
                d = read_csv(f'out/work/primary/glm/singlefactor/{args.outlier}/{cat}/{tissue}/{cond}/{comp}.csv',index_col=0)
                d['Category'] = cat
                d['Tissue'] = tissue
                d['Condition'] = cond
                d['Comparison'] = comp
                for m in d.index:
                    if d.loc[m,'Pr(>|z|)'] < args.level:
                        sig[m,tissue,cond,comp] = True
                    else:
                        sig[m,tissue,cond,comp] = False
                    if d.loc[m,'Estimate'] > 0.:
                        up[m,tissue,cond,comp] = True
                    else:
                        up[m,tissue,cond,comp] = False
                datasets.append(d)
data = concat(datasets,axis=0).dropna()

wb = Workbook()
ws = wb.active
ws.title = 'primary-crosspop'

# metabolite name column
ws.freeze_panes = 'A4'
ws.column_dimensions['A'].width = 5.
ws.column_dimensions['B'].width = 25.
for k in range(27):
    ws.column_dimensions[get_column_letter(k+3)].width = 4.
for i,comp in zip(range(len(comparisons)),comparisons):
    c = ws.cell(row=1, column=3+i*9,)
    c.value = ' vs. '.join(comparisons[comp])
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1,start_column=3+i*9,end_row=1,end_column=2+(i+1)*9-1)
    for j,tissue in zip(range(len(tissues)),tissues):
        c = ws.cell(row=2, column=3+i*9+j*3,)
        c.value = tissue
        c.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=2,start_column=3+i*9+j*3,end_row=2,end_column=3+i*9+(j+1)*3-1)
        for k,cond in zip(range(len(condmap)),condmap):
            c = ws.cell(row=3, column=3+i*9+j*3+k,)
            c.value = cond
            c.alignment = Alignment(horizontal="center")

            cat_start = None
            l = 4
            for cat in categories:
                if cat_start != None:
                    ws.merge_cells(start_row=cat_start,end_row=l-1,start_column=1,end_column=1)
                ws.cell(row=l,column=1).value = categories[cat]
                ws.cell(row=l,column=1).alignment = Alignment(text_rotation=90.,vertical="center")
                cat_start = l
                d = data.loc[(data['Category'] == cat) & (data['Tissue'] == tissue) & (data['Condition'] == cond) & (data['Comparison'] == comp)]
                for m in sorted(d.index):
                    ws.cell(row=l,column=2).value = m
                    # shade muscle
                    if tissue == 'Muscle':
                        c = ws.cell(row=l,column=3+i*9+j*3+k,)
                        c.fill = PatternFill('solid','F0F0F0F0')
                    if sig[m,tissue,cond,comp]:
                        c = ws.cell(row=l,column=3+i*9+j*3+k,)
                        if up[m,tissue,cond,comp]:
                            c.value = '↑'
                            c.alignment = Alignment(horizontal="center",vertical="center")
                        else:
                            c.value = '↓'
                            c.alignment = Alignment(horizontal="center",vertical="center")
                        if comp == 'PvS' or comp == 'TvS':
                            if up[m,tissue,cond,'PvS'] == up[m,tissue,cond,'TvS']:
                                if (not sig[m,tissue,cond,'PvT']) or (sig[m,tissue,cond,'PvS'] == sig[m,tissue,cond,'TvS']):
                                    #c.value = 'xx'
                                    c.fill = PatternFill('solid','00CDCDFF')
                    if (i*9+j*3+k+1) % 9 == 0 and i < 2:
                        c = ws.cell(row=l,column=3+i*9+j*3+k,)
                        c.border = Border(right=Side(border_style='double',color='00000000'))
                    l += 1
            ws.merge_cells(start_row=cat_start,end_row=l-1,start_column=1,end_column=1)

wb.save(args.output)
